"""
Neo4j Phase Profiler - Detailed Performance Analysis
====================================================
Tracks detailed timing for each phase of Neo4j GraphRAG operations:
- Connection establishment
- Embedding generation
- Vector search (similarity)
- Fulltext/keyword search
- Hybrid scoring & ranking
- Graph traversal/expansion
- Result formatting

Features:
- Per-phase timing breakdown
- Concurrent load testing with phase tracking
- Phase-level bottleneck identification
- Statistical analysis per phase
- Excel reports with phase comparison charts
"""

import requests
import time
import json
import threading
import concurrent.futures
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd
from dataclasses import dataclass, asdict
import statistics
from pathlib import Path
import random
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv(Path(__file__).parent.parent / '.env')

# Rich console (optional)
try:
    from rich.console import Console
    from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn
    from rich.table import Table
    from rich.live import Live
    RICH_AVAILABLE = True
    console = Console()
except ImportError:
    RICH_AVAILABLE = False
    console = None

# Excel styling imports
try:
    from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
    EXCEL_STYLING_AVAILABLE = True
except ImportError:
    EXCEL_STYLING_AVAILABLE = False
    # Define dummy classes to prevent errors
    class NamedStyle:
        def __init__(self, name): pass
    class Font:
        def __init__(self, **kwargs): pass
    class PatternFill:
        def __init__(self, **kwargs): pass
    class Alignment:
        def __init__(self, **kwargs): pass
    class Border:
        def __init__(self, **kwargs): pass
    class Side:
        def __init__(self, **kwargs): pass

# System metrics (optional)
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# ========================================
# CONFIGURATION
# ========================================

class Config:
    """Test configuration"""
    API_URL = os.getenv("API_URL")
    QUEUE_STATUS_URL = os.getenv("QUEUE_STATUS_URL")
    
    OUTPUT_DIR = Path("test_results")
    
    # Load test parameters
    MIN_USERS = 1
    MAX_USERS = 10
    USERS_INCREMENT = 2
    WAIT_BETWEEN_STEPS = 5
    
    REQUEST_TIMEOUT = None  # No timeout - wait indefinitely
    MAX_RETRIES = 2
    RETRY_DELAY = 2
    
    # Phase timing expectations (ms) - used for warnings
    EXPECTED_EMBED_TIME = 500
    EXPECTED_VECTOR_SEARCH_TIME = 200
    EXPECTED_FULLTEXT_SEARCH_TIME = 150
    EXPECTED_GRAPH_TRAVERSE_TIME = 300


# Test questions
QUESTIONS = [
    "Could you please explain how the mining sector is expected to contribute to inclusivity, employment opportunities, and the development of rural communities in Sarawak?",
    "Can you explain how the handling daily operation of the facility is planned for tourism facilities development?",
    "Can u tell me how state-owned universties help with UP-DLP Sarawak?",
    "Apakah tindakan Jabatan Keselamatan dan Kesihatan Pekerjaan terhadap syarikat yang terlibat dalam bahaya utama di SIP?",
    "Boleh jelaskan berapa projek yang telah dilaksanakan oleh Sr Aman Development Agency (SADA) dan jumlah peruntukan yang diterima untuk tahun 2025?",
    "Can you tell me what are the plans for Totally Protected Areas facilities upgrade?",
    "Could you please explain the main strategies and initiatives outlined for the tourism sector in Sarawak?",
    "Can you explain how the Rajah Brooke dynasty influenced the cultural and historical development of the Sarawak Delta Geopark?",
    "Can you explain the repayment terms for an advance to purchase a new vehicle based on the Sarawak General Order 1996?",
    "Can you provide detailed information on how the projects under the tourism sector ensure timely completion with minimal delays?",
    "Datuk Seri Alexander Nanta Linggi tu dia kerja apa dalam Kabinet Malaysia sekarang?",
    "Can you explain in detail the strategies and expected outcomes of the Sarawak Heritage Ordinance administration?",
    "What are the key details and benefits of the Sarawak tourism promotion incentives?",
    "Can you provide details on the initiative for Securing Business Events for Miri?",
    "How is Sarawak planning to enhance food production for export?",
    "What are the key targets and expected economic impacts of the Business Events 2021 to 2025 initiatives in Sarawak?",
    "Could you please explain how the manufacturing sector's initiatives will benefit rural communities in Sarawak?",
    "Can you tell me what facilities will be developed at Limbang Mangrove National Park?",
    "How can I obtain permission for use of content from the PCDS 2030 Highlights 2023 document?",
    "Who is the State Secretary of Sarawak based on Cabinet Members of Malaysia and Sarawak Government?"
]


# ========================================
# DATA MODELS
# ========================================

@dataclass
class PhaseTimings:
    """Detailed timing breakdown for a single request"""
    # Network & connection
    connection_time_ms: Optional[float] = None
    dns_lookup_ms: Optional[float] = None
    tcp_connect_ms: Optional[float] = None
    
    # API-reported timings (from response)
    embed_time_ms: Optional[float] = None
    # Optional embedding sub-phases (if server provides)
    embed_queue_wait_ms: Optional[float] = None
    embed_tokenize_ms: Optional[float] = None
    embed_forward_ms: Optional[float] = None
    embed_postprocess_ms: Optional[float] = None
    hybrid_time_ms: Optional[float] = None
    mmr_time_ms: Optional[float] = None
    cross_doc_time_ms: Optional[float] = None
    graph_traverse_time_ms: Optional[float] = None
    format_context_time_ms: Optional[float] = None
    # Optional Neo4j sub-phases
    neo4j_seed_fetch_ms: Optional[float] = None
    neo4j_traverse_exec_ms: Optional[float] = None
    
    # Client-side measurements
    total_request_time_ms: Optional[float] = None
    response_transfer_time_ms: Optional[float] = None
    client_json_parse_time_ms: Optional[float] = None
    response_bytes: Optional[int] = None
    
    # Optional server-reported finer phases
    api_inbound_parse_time_ms: Optional[float] = None
    api_json_serialize_time_ms: Optional[float] = None
    api_network_send_time_ms: Optional[float] = None


@dataclass
class RequestResult:
    """Result of a single request with phase timings"""
    timestamp: str
    step: int
    user_index: int
    question: str
    success: bool
    status_code: Optional[int]
    error_message: str
    retry_count: int
    
    # Phase timings
    phases: PhaseTimings
    
    # Metadata
    seeds_count: int = 0
    facts_length: int = 0
    answer: str = ""
    
    # Embedding metrics
    concurrent_embeds: int = 0  # Number of concurrent embedding requests in this step

    # Application/transport indicators (defaulted at end to satisfy dataclass ordering)
    app_success: bool = False
    transport_warning: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        result = asdict(self)
        # Flatten phases for Excel
        phases_dict = result.pop('phases')
        result.update({f'phase_{k}': v for k, v in phases_dict.items()})
        return result


@dataclass
class StepMetrics:
    """Aggregated metrics for a load step"""
    step: int
    concurrent_users: int
    total_requests: int
    successful_requests: int
    failed_requests: int
    success_rate: float
    
    # Average phase timings
    avg_connection_ms: float
    avg_embed_ms: float
    avg_hybrid_ms: float
    avg_mmr_ms: float
    avg_cross_doc_ms: float
    avg_graph_traverse_ms: float
    avg_format_context_ms: float
    avg_total_request_ms: float
    avg_client_json_parse_ms: float
    avg_response_bytes: float
    
    # Percentiles (P95)
    p95_connection_ms: float
    p95_embed_ms: float
    p95_hybrid_ms: float
    p95_mmr_ms: float
    p95_cross_doc_ms: float
    p95_graph_traverse_ms: float
    p95_total_request_ms: float
    p95_client_json_parse_ms: float
    
    # Bottleneck identification
    slowest_phase: str
    slowest_phase_avg_ms: float
    slowest_phase_pct: float
    cause_label: str
    top_processing_contributors: str
    
    # Throughput metrics
    throughput_rps: float
    
    # Request capacity metrics
    request_throughput: float  # Requests per second (wall-clock)
    avg_embeddings_per_request: float
    total_embedding_requests: int
    # Host metrics (optional)
    avg_cpu_percent: float = 0.0
    avg_cpu_iowait_percent: float = 0.0
    avg_mem_percent: float = 0.0
    avg_swap_percent: float = 0.0
    load1_per_core_avg: float = 0.0
    net_rx_mb_s: float = 0.0
    net_tx_mb_s: float = 0.0
    disk_read_mb_s: float = 0.0
    disk_write_mb_s: float = 0.0
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


# ========================================
# API CLIENT WITH PHASE TRACKING
# ========================================

class GraphRAGClient:
    """Client with detailed phase timing tracking"""
    
    def __init__(self, config: Config):
        self.config = config
        self.session = requests.Session()
        
    def send_request(self, question: str, step: int, user_index: int) -> RequestResult:
        """Send request and track all phase timings"""
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S%z")
        phases = PhaseTimings()
        
        result = RequestResult(
            timestamp=timestamp,
            step=step,
            user_index=user_index,
            question=question[:100],
            success=False,
            status_code=None,
            error_message="",
            retry_count=0,
            phases=phases
        )
        
        headers = {"Content-Type": "application/json"}
        payload = {"question": question}
        
        for attempt in range(self.config.MAX_RETRIES):
            try:
                # PHASE 1: Measure total request time
                request_start = time.perf_counter()
                
                response = self.session.post(
                    self.config.API_URL,
                    headers=headers,
                    json=payload,
                    timeout=self.config.REQUEST_TIMEOUT
                )
                
                request_end = time.perf_counter()
                phases.total_request_time_ms = (request_end - request_start) * 1000
                
                result.status_code = response.status_code
                result.retry_count = attempt
                
                # Measure transfer and parse times
                try:
                    content_start = time.perf_counter()
                    content = response.content
                    phases.response_transfer_time_ms = (time.perf_counter() - content_start) * 1000
                    phases.response_bytes = len(content) if content is not None else 0
                except Exception:
                    content = None
                
                try:
                    parse_start = time.perf_counter()
                    data = response.json() if content is None else json.loads(content.decode('utf-8', errors='replace'))
                    phases.client_json_parse_time_ms = (time.perf_counter() - parse_start) * 1000
                except Exception as parse_error:
                    data = None
                    # Debug: Print what we actually received when JSON parsing fails
                    if attempt == 0:  # Only print on first attempt
                        print(f"[DEBUG] JSON parse failed: {type(parse_error).__name__}: {str(parse_error)[:100]}")
                        if content:
                            content_preview = content.decode('utf-8', errors='replace')[:500]
                            print(f"[DEBUG] Response content preview: {content_preview}")
                            print(f"[DEBUG] Response length: {len(content)} bytes")
                        else:
                            print(f"[DEBUG] Response content is None/empty")
                
                # Determine success based on HTTP status AND response structure
                # If API returns success field, use it; otherwise check HTTP status + valid JSON
                api_success = False
                if isinstance(data, dict):
                    if "success" in data:
                        # API explicitly indicates success/failure
                        api_success = data.get("success", False)
                        
                        # Debug: Print first failure details
                        if not api_success and attempt == 0:
                            print(f"[DEBUG] API returned success=false. Response keys: {list(data.keys())}")
                            if "message" in data:
                                print(f"[DEBUG] Message: {data['message'][:200]}")
                            if "error_type" in data:
                                print(f"[DEBUG] Error type: {data['error_type']}")
                    elif response.status_code == 200:
                        # No explicit success field, but HTTP 200 with valid JSON = success
                        api_success = True
                    
                result.app_success = api_success
                result.success = api_success  # Use application-level success as primary metric
                
                if response.status_code and response.status_code >= 500 and api_success:
                    result.transport_warning = f"http_{response.status_code}_with_success"
                
                if isinstance(data, dict) and api_success:
                    # PHASE 2: Extract API-reported timings
                    timings = data.get("timings", {})
                    if timings:
                        phases.embed_time_ms = timings.get("embed", 0) * 1000
                        # Optional embedding sub-phases
                        phases.embed_queue_wait_ms = (timings.get("embed_queue_wait", 0) or 0) * 1000
                        phases.embed_tokenize_ms = (timings.get("embed_tokenize", 0) or 0) * 1000
                        phases.embed_forward_ms = (timings.get("embed_forward", 0) or 0) * 1000
                        phases.embed_postprocess_ms = (timings.get("embed_postprocess", 0) or 0) * 1000
                        phases.hybrid_time_ms = timings.get("hybrid", 0) * 1000
                        phases.mmr_time_ms = timings.get("mmr", 0) * 1000
                        phases.cross_doc_time_ms = timings.get("cross_doc", 0) * 1000
                        phases.graph_traverse_time_ms = timings.get("graph_traverse", 0) * 1000
                        phases.format_context_time_ms = timings.get("format_context", 0) * 1000
                        # Optional Neo4j sub-phases
                        phases.neo4j_seed_fetch_ms = (timings.get("neo4j_seed_fetch", 0) or 0) * 1000
                        phases.neo4j_traverse_exec_ms = (timings.get("neo4j_traverse_exec", 0) or 0) * 1000
                        # Optional finer API timings if provided by server
                        phases.api_inbound_parse_time_ms = (timings.get("inbound_parse", 0) or 0) * 1000
                        phases.api_json_serialize_time_ms = (timings.get("json_serialize", 0) or 0) * 1000
                        phases.api_network_send_time_ms = (timings.get("network_send", 0) or 0) * 1000
                    
                    # Extract metadata
                    result.seeds_count = len(data.get("seeds", []))
                    result.facts_length = len(data.get("facts", ""))
                    result.answer = data.get("answer", "")
                    
                    return result
                else:
                    # Application reported failure or JSON invalid
                    if isinstance(data, dict) and not api_success:
                        api_msg = data.get("message", "")
                        error_type = data.get("error_type", "")
                        error_details = data.get("error_details", "")
                        
                        # Build comprehensive error message
                        parts = [f"HTTP {response.status_code}"]
                        if api_msg:
                            parts.append(api_msg)
                        if error_type:
                            parts.append(f"[{error_type}]")
                        if error_details and error_details != api_msg:
                            parts.append(f"({error_details[:100]})")
                        
                        result.error_message = ": ".join(parts) if len(parts) > 1 else parts[0]
                    else:
                        # Non-JSON or other HTTP error without JSON
                        if content:
                            content_type = response.headers.get('content-type', 'unknown')
                            result.error_message = f"HTTP {response.status_code}: Invalid JSON (content-type: {content_type}, size: {len(content)})"
                        else:
                            result.error_message = f"HTTP {response.status_code}: Empty response body"
                    
                    # Retry logic
                    if attempt < self.config.MAX_RETRIES - 1:
                        time.sleep(self.config.RETRY_DELAY * (attempt + 1))
                        continue
                    else:
                        return result
                    
            except requests.exceptions.Timeout:
                result.error_message = f"Timeout after {self.config.REQUEST_TIMEOUT}s"
                result.retry_count = attempt
                if attempt < self.config.MAX_RETRIES - 1:
                    time.sleep(self.config.RETRY_DELAY * (attempt + 1))
                    continue
                return result
                
            except Exception as e:
                result.error_message = f"Error: {str(e)[:100]}"
                result.retry_count = attempt
                return result
        
        return result
    
    def close(self):
        """Close session"""
        self.session.close()


# ========================================
# TEST EXECUTOR
# ========================================

class PhaseProfiler:
    """Execute tests and analyze phase timings"""
    
    def __init__(self, config: Config):
        self.config = config
        self.client = GraphRAGClient(config)
        self.all_results: List[RequestResult] = []
        self.step_metrics: List[StepMetrics] = []
        
    def run_concurrent_requests(self, num_users: int, step: int) -> List[RequestResult]:
        """Run concurrent requests"""
        results = []
        
        if RICH_AVAILABLE:
            console.print(f"[dim]Testing {num_users} concurrent users...[/dim]")
        else:
            print(f"Testing {num_users} concurrent users...")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=num_users) as executor:
            questions = [random.choice(QUESTIONS) for _ in range(num_users)]
            futures = [
                executor.submit(self.client.send_request, q, step, i)
                for i, q in enumerate(questions)
            ]
            
            for future in concurrent.futures.as_completed(futures):
                try:
                    result = future.result()
                    result.concurrent_embeds = num_users  # Track concurrent load
                    results.append(result)
                except Exception as e:
                    if RICH_AVAILABLE:
                        console.print(f"[red]Error: {e}[/red]")
                    else:
                        print(f"Error: {e}")
        
        # Report failures if any
        failures = [r for r in results if not r.success]
        if failures:
            if RICH_AVAILABLE:
                console.print(f"[yellow]⚠️  {len(failures)}/{len(results)} requests failed[/yellow]")
                # Show first few unique error messages
                unique_errors = {}
                for f in failures:
                    msg = f.error_message or "Unknown error"
                    unique_errors[msg] = unique_errors.get(msg, 0) + 1
                for msg, count in list(unique_errors.items())[:3]:
                    console.print(f"[dim]  • {msg} ({count}x)[/dim]")
            else:
                print(f"⚠️  {len(failures)}/{len(results)} requests failed")
                unique_errors = {}
                for f in failures:
                    msg = f.error_message or "Unknown error"
                    unique_errors[msg] = unique_errors.get(msg, 0) + 1
                for msg, count in list(unique_errors.items())[:3]:
                    print(f"  • {msg} ({count}x)")
        
        return results
    
    def calculate_step_metrics(self, results: List[RequestResult], step: int, concurrent_users: int, step_duration_sec: float = 0) -> StepMetrics:
        """Calculate aggregated phase metrics"""
        total = len(results)
        successful = sum(1 for r in results if r.success)
        success_rate = (successful / total * 100) if total > 0 else 0
        
        # Extract phase timings from successful requests
        def safe_values(attr: str) -> List[float]:
            values = []
            for r in results:
                if r.success:
                    val = getattr(r.phases, attr, None)
                    if val is not None and val > 0:
                        values.append(val)
            return values
        
        embed_times = safe_values('embed_time_ms')
        embed_queue_times = safe_values('embed_queue_wait_ms')
        embed_tokenize_times = safe_values('embed_tokenize_ms')
        embed_forward_times = safe_values('embed_forward_ms')
        embed_post_times = safe_values('embed_postprocess_ms')
        hybrid_times = safe_values('hybrid_time_ms')
        mmr_times = safe_values('mmr_time_ms')
        cross_doc_times = safe_values('cross_doc_time_ms')
        traverse_times = safe_values('graph_traverse_time_ms')
        neo4j_seed_fetch_times = safe_values('neo4j_seed_fetch_ms')
        neo4j_traverse_exec_times = safe_values('neo4j_traverse_exec_ms')
        format_times = safe_values('format_context_time_ms')
        total_times = safe_values('total_request_time_ms')
        json_parse_times = safe_values('client_json_parse_time_ms')
        response_sizes = safe_values('response_bytes')
        
        def avg(vals): return statistics.mean(vals) if vals else 0
        def p95(vals): return statistics.quantiles(vals, n=20)[18] if len(vals) > 1 else (vals[0] if vals else 0)
        
        # Identify bottleneck among raw phase timings only
        phase_avgs = {
            "embedding": avg(embed_times),
            "hybrid_search": avg(hybrid_times),
            "mmr": avg(mmr_times),
            "cross_doc": avg(cross_doc_times),
            "graph_traverse": avg(traverse_times),
            "format_context": avg(format_times),
        }
        slowest_phase = max(phase_avgs, key=phase_avgs.get) if phase_avgs else "unknown"
        slowest_phase_avg = phase_avgs.get(slowest_phase, 0)
        avg_total_ms = avg(total_times)
        slowest_phase_pct = (slowest_phase_avg / avg_total_ms * 100) if avg_total_ms > 0 else 0

        # Classify likely cause
        def classify_bottleneck(key: str) -> str:
            key = key or ""
            if key == "embedding":
                return "embedding_service"
            if key in ("graph_traverse", "hybrid_search"):
                return "neo4j_database"
            if key in ("format_context", "mmr", "cross_doc"):
                return "application_processing"
            return "unknown"

        cause_label = classify_bottleneck(slowest_phase)

        # Top contributors by time
        contrib = [
            ("Embedding", avg(embed_times)),
            ("Hybrid", avg(hybrid_times)),
            ("MMR", avg(mmr_times)),
            ("CrossDoc", avg(cross_doc_times)),
            ("Traverse", avg(traverse_times)),
            ("Format", avg(format_times)),
        ]
        contrib = [(name, val, (val / avg_total_ms * 100) if avg_total_ms > 0 else 0) for name, val in contrib]
        contrib.sort(key=lambda x: x[1], reverse=True)
        top_processing_contributors = ", ".join([f"{name} {pct:.0f}%" for name, _, pct in contrib if pct > 0][:3])
        
        # Calculate throughput (requests per second)
        # Use the step duration from the calling method, but since we don't have it here,
        # we'll calculate it based on concurrent users and average response time
        # Throughput = concurrent_users / (avg_response_time / 1000)
        avg_response_time_sec = avg_total_ms / 1000 if total_times else 0
        throughput_rps = concurrent_users / avg_response_time_sec if avg_response_time_sec > 0 else 0
        
        # Calculate request throughput
        # Each successful request does 1 embedding call
        # Real throughput = total requests completed / wall-clock time
        total_embedding_requests = successful  # 1 embedding per successful request
        request_throughput = successful / step_duration_sec if step_duration_sec > 0 else 0
        
        return StepMetrics(
            step=step,
            concurrent_users=concurrent_users,
            total_requests=total,
            successful_requests=successful,
            failed_requests=total - successful,
            success_rate=success_rate,
            
            avg_connection_ms=0,  # Not tracked in current API
            avg_embed_ms=avg(embed_times),
            avg_hybrid_ms=avg(hybrid_times),
            avg_mmr_ms=avg(mmr_times),
            avg_cross_doc_ms=avg(cross_doc_times),
            avg_graph_traverse_ms=avg(traverse_times),
            avg_format_context_ms=avg(format_times),
            avg_total_request_ms=avg(total_times),
            avg_client_json_parse_ms=avg(json_parse_times),
            avg_response_bytes=avg(response_sizes),
            
            p95_connection_ms=0,
            p95_embed_ms=p95(embed_times),
            p95_hybrid_ms=p95(hybrid_times),
            p95_mmr_ms=p95(mmr_times),
            p95_cross_doc_ms=p95(cross_doc_times),
            p95_graph_traverse_ms=p95(traverse_times),
            p95_total_request_ms=p95(total_times),
            p95_client_json_parse_ms=p95(json_parse_times),

            slowest_phase=slowest_phase,
            slowest_phase_avg_ms=slowest_phase_avg,
            slowest_phase_pct=slowest_phase_pct,
            cause_label=cause_label,
            top_processing_contributors=top_processing_contributors,
            throughput_rps=throughput_rps,

            request_throughput=request_throughput,
            avg_embeddings_per_request=1.0,  # Currently 1 embedding per request
            total_embedding_requests=total_embedding_requests,
            avg_cpu_percent=getattr(self, "_last_step_cpu_avg", 0.0),
            avg_cpu_iowait_percent=getattr(self, "_last_step_cpu_iowait_avg", 0.0),
            avg_mem_percent=getattr(self, "_last_step_mem_avg", 0.0),
            avg_swap_percent=getattr(self, "_last_step_swap_avg", 0.0),
            load1_per_core_avg=getattr(self, "_last_step_load1_per_core_avg", 0.0),
            net_rx_mb_s=getattr(self, "_last_step_net_rx_mb_s", 0.0),
            net_tx_mb_s=getattr(self, "_last_step_net_tx_mb_s", 0.0),
            disk_read_mb_s=getattr(self, "_last_step_disk_read_mb_s", 0.0),
            disk_write_mb_s=getattr(self, "_last_step_disk_write_mb_s", 0.0),
        )
    
    def run_test(self):
        """Run the phase profiling test"""
        if RICH_AVAILABLE:
            console.print("\n[bold cyan]🔍 Neo4j Phase Profiler - Starting Test[/bold cyan]\n")
        else:
            print("\n🔍 Neo4j Phase Profiler - Starting Test\n")
        
        step = 0
        try:
            for num_users in range(self.config.MIN_USERS, self.config.MAX_USERS + 1, self.config.USERS_INCREMENT):
                step += 1
                
                if RICH_AVAILABLE:
                    console.print(f"\n[bold yellow]📊 Step {step}: {num_users} concurrent users[/bold yellow]")
                else:
                    print(f"\n📊 Step {step}: {num_users} concurrent users")
                
                # Start system monitor (optional)
                cpu_samples = []
                iowait_samples = []
                mem_samples = []
                swap_samples = []
                load1_samples = []
                stop_event = threading.Event()

                # Baselines for IO counters
                if PSUTIL_AVAILABLE:
                    try:
                        disk_io_start = psutil.disk_io_counters()
                        net_io_start = psutil.net_io_counters()
                    except Exception:
                        disk_io_start = None
                        net_io_start = None

                def sys_monitor():
                    if not PSUTIL_AVAILABLE:
                        return
                    while not stop_event.is_set():
                        try:
                            # Use cpu_times_percent with interval to also capture iowait
                            t = psutil.cpu_times_percent(interval=0.5)
                            cpu_percent = 100.0 - getattr(t, 'idle', 0.0)
                            cpu_samples.append(cpu_percent)
                            iowait_samples.append(getattr(t, 'iowait', 0.0) or 0.0)

                            vm = psutil.virtual_memory()
                            mem_samples.append(getattr(vm, 'percent', 0.0) or 0.0)
                            sw = psutil.swap_memory()
                            swap_samples.append(getattr(sw, 'percent', 0.0) or 0.0)

                            try:
                                la1, la5, la15 = os.getloadavg()
                                cores = psutil.cpu_count(logical=True) or 1
                                load1_samples.append(la1 / cores)
                            except Exception:
                                pass
                        except Exception:
                            break

                monitor_thread = None
                if PSUTIL_AVAILABLE:
                    monitor_thread = threading.Thread(target=sys_monitor, daemon=True)
                    monitor_thread.start()

                step_start = time.perf_counter()
                step_results = self.run_concurrent_requests(num_users, step)
                step_duration = time.perf_counter() - step_start

                # Stop system monitor and compute averages/deltas
                if PSUTIL_AVAILABLE:
                    stop_event.set()
                    if monitor_thread:
                        monitor_thread.join(timeout=1)
                    self._last_step_cpu_avg = statistics.mean(cpu_samples) if cpu_samples else 0.0
                    self._last_step_cpu_iowait_avg = statistics.mean(iowait_samples) if iowait_samples else 0.0
                    self._last_step_mem_avg = statistics.mean(mem_samples) if mem_samples else 0.0
                    self._last_step_swap_avg = statistics.mean(swap_samples) if swap_samples else 0.0
                    self._last_step_load1_per_core_avg = statistics.mean(load1_samples) if load1_samples else 0.0

                    # Compute IO bandwidths using deltas
                    try:
                        disk_io_end = psutil.disk_io_counters()
                        net_io_end = psutil.net_io_counters()
                    except Exception:
                        disk_io_end = None
                        net_io_end = None

                    dur = step_duration if step_duration > 0 else 1e-6
                    if disk_io_start and disk_io_end:
                        read_mb = max(0.0, (disk_io_end.read_bytes - disk_io_start.read_bytes) / (1024 * 1024))
                        write_mb = max(0.0, (disk_io_end.write_bytes - disk_io_start.write_bytes) / (1024 * 1024))
                        self._last_step_disk_read_mb_s = read_mb / dur
                        self._last_step_disk_write_mb_s = write_mb / dur
                    else:
                        self._last_step_disk_read_mb_s = 0.0
                        self._last_step_disk_write_mb_s = 0.0

                    if net_io_start and net_io_end:
                        rx_mb = max(0.0, (net_io_end.bytes_recv - net_io_start.bytes_recv) / (1024 * 1024))
                        tx_mb = max(0.0, (net_io_end.bytes_sent - net_io_start.bytes_sent) / (1024 * 1024))
                        self._last_step_net_rx_mb_s = rx_mb / dur
                        self._last_step_net_tx_mb_s = tx_mb / dur
                    else:
                        self._last_step_net_rx_mb_s = 0.0
                        self._last_step_net_tx_mb_s = 0.0
                else:
                    self._last_step_cpu_avg = 0.0
                    self._last_step_cpu_iowait_avg = 0.0
                    self._last_step_mem_avg = 0.0
                    self._last_step_swap_avg = 0.0
                    self._last_step_load1_per_core_avg = 0.0
                    self._last_step_disk_read_mb_s = 0.0
                    self._last_step_disk_write_mb_s = 0.0
                    self._last_step_net_rx_mb_s = 0.0
                    self._last_step_net_tx_mb_s = 0.0
                
                metrics = self.calculate_step_metrics(step_results, step, num_users, step_duration)
                self.all_results.extend(step_results)
                self.step_metrics.append(metrics)
                
                self._display_step_results(metrics)

                # Check for failed requests and stop if any
                failed_results = [r for r in step_results if not r.success]
                if failed_results:
                    if RICH_AVAILABLE:
                        console.print(f"[red]❌ Failed request detected ({len(failed_results)}/{len(step_results)} failed). Stopping test.[/red]")
                        for fr in failed_results:
                            console.print(f"[red]   Error: {fr.error_message} | Status: {fr.status_code}[/red]")
                    else:
                        print(f"❌ Failed request detected ({len(failed_results)}/{len(step_results)} failed). Stopping test.")
                        for fr in failed_results:
                            print(f"   Error: {fr.error_message} | Status: {fr.status_code}")
                    break
                
                if num_users < self.config.MAX_USERS:
                    time.sleep(self.config.WAIT_BETWEEN_STEPS)
        
        except KeyboardInterrupt:
            if RICH_AVAILABLE:
                console.print("\n[yellow]⚠️  Test interrupted[/yellow]")
            else:
                print("\n⚠️  Test interrupted")
        
        finally:
            self.client.close()
        
        self._save_results()
        self._display_summary()
    
    def _display_step_results(self, m: StepMetrics):
        """Display phase breakdown for a step"""
        if RICH_AVAILABLE:
            table = Table(title="Phase Timing Breakdown", show_header=True)
            table.add_column("Phase", style="cyan")
            table.add_column("Avg (ms)", justify="right")
            table.add_column("P95 (ms)", justify="right")
            table.add_column("% of Total", justify="right")
            
            # Use total request time as the base for percentage calculations
            total = m.avg_total_request_ms if m.avg_total_request_ms > 0 else 1
            
            # Raw phase timings only
            phases = [
                ("Embedding", m.avg_embed_ms, m.p95_embed_ms),
                ("Hybrid Search", m.avg_hybrid_ms, m.p95_hybrid_ms),
                ("MMR", m.avg_mmr_ms, m.p95_mmr_ms),
                ("Cross Doc", m.avg_cross_doc_ms, m.p95_cross_doc_ms),
                ("Graph Traverse", m.avg_graph_traverse_ms, m.p95_graph_traverse_ms),
                ("Format Context", m.avg_format_context_ms, 0),
            ]
            
            # Calculate percentages for all phases and track sum
            phase_percentages = []
            percentage_sum = 0.0
            
            for name, avg_val, p95_val in phases:
                pct = (avg_val / total * 100) if total > 0 else 0
                rounded_pct = round(pct, 1)
                phase_percentages.append((name, avg_val, p95_val, rounded_pct))
                percentage_sum += rounded_pct
            
            # Calculate overhead percentage to make total exactly 100%
            overhead_ms = total - sum([avg_val for _, avg_val, _ in phases])
            overhead_pct = round(100.0 - percentage_sum, 1)  # Force to 100%
            
            for name, avg_val, p95_val, pct in phase_percentages:
                # Normalize key names for comparison
                key_name = name.lower().replace(" ", "_")
                style = "bold red" if key_name == m.slowest_phase else ""
                table.add_row(
                    name,
                    f"{avg_val:.1f}",
                    f"{p95_val:.1f}" if p95_val > 0 else "-",
                    f"{pct:.1f}%",
                    style=style
                )
            
            # Add overhead row with adjusted percentage
            table.add_row(
                "Other/Overhead",
                f"{overhead_ms:.1f}",
                "-",
                f"{overhead_pct:.1f}%",
                style="dim"
            )
            
            # Add separator and total
            table.add_row("", "", "", "", style="dim")
            table.add_row(
                "TOTAL REQUEST",
                f"{m.avg_total_request_ms:.1f}",
                f"{m.p95_total_request_ms:.1f}",
                "100%",
                style="bold"
            )
            
            console.print(table)
            console.print(f"[yellow]⚠️  Bottleneck: {m.slowest_phase} ({m.slowest_phase_avg_ms:.1f}ms, {m.slowest_phase_pct:.1f}% of total)[/yellow]")
            console.print(f"Success Rate: [green]{m.success_rate:.1f}%[/green]")
            console.print(f"[cyan]📊 Request Throughput: {m.request_throughput:.2f} req/sec[/cyan] | "
                         f"[dim]Avg embed time: {m.avg_embed_ms:.0f}ms (concurrent: {m.concurrent_users})[/dim]")
            if m.top_processing_contributors:
                console.print(f"[dim]🔎 Top contributors: {m.top_processing_contributors}[/dim]")
            console.print(f"[dim]📦 Client JSON parse avg: {m.avg_client_json_parse_ms:.1f}ms | Resp size avg: {m.avg_response_bytes/1024:.1f} KB[/dim]")
            # If embedding is bottleneck, attempt sub-phase breakdown from available per-request data
            if m.slowest_phase == "embedding":
                eq, et, ef, ep = [], [], [], []
                for r in self.all_results:
                    if r.success and r.step == m.step:
                        p = r.phases
                        if p.embed_queue_wait_ms: eq.append(p.embed_queue_wait_ms)
                        if p.embed_tokenize_ms: et.append(p.embed_tokenize_ms)
                        if p.embed_forward_ms: ef.append(p.embed_forward_ms)
                        if p.embed_postprocess_ms: ep.append(p.embed_postprocess_ms)
                total_embed = (statistics.mean(eq) if eq else 0) + (statistics.mean(et) if et else 0) + (statistics.mean(ef) if ef else 0) + (statistics.mean(ep) if ep else 0)
                if total_embed > 0:
                    def share(vals):
                        v = statistics.mean(vals) if vals else 0
                        return (v / total_embed * 100) if total_embed > 0 else 0
                    console.print(f"[dim]🧩 Embedding breakdown: Queue {share(eq):.0f}%, Tokenize {share(et):.0f}%, Forward {share(ef):.0f}%, Post {share(ep):.0f}%[/dim]")
            if m.avg_cpu_percent:
                console.print(f"[dim]🖥️  Host CPU avg during step: {m.avg_cpu_percent:.1f}%[/dim]")
            if m.avg_mem_percent or m.avg_cpu_iowait_percent:
                console.print(f"[dim]🧠 Mem avg: {m.avg_mem_percent:.1f}% | ⏳ iowait: {m.avg_cpu_iowait_percent:.1f}% | load1/core: {m.load1_per_core_avg:.2f}[/dim]")
            if m.net_rx_mb_s or m.net_tx_mb_s or m.disk_read_mb_s or m.disk_write_mb_s:
                console.print(f"[dim]🌐 Net RX/TX: {m.net_rx_mb_s:.2f}/{m.net_tx_mb_s:.2f} MB/s | 💾 Disk R/W: {m.disk_read_mb_s:.2f}/{m.disk_write_mb_s:.2f} MB/s[/dim]")
        else:
            print(f"  Avg Total: {m.avg_total_request_ms:.1f}ms")
            print(f"  - Embedding: {m.avg_embed_ms:.1f}ms")
            print(f"  - Hybrid Search: {m.avg_hybrid_ms:.1f}ms")
            print(f"  - MMR: {m.avg_mmr_ms:.1f}ms")
            print(f"  - Cross Doc: {m.avg_cross_doc_ms:.1f}ms")
            print(f"  - Graph Traverse: {m.avg_graph_traverse_ms:.1f}ms")
            print(f"  - Format Context: {m.avg_format_context_ms:.1f}ms")
            print(f"  Bottleneck: {m.slowest_phase} ({m.slowest_phase_avg_ms:.1f}ms)")
            print(f"  Success Rate: {m.success_rate:.1f}%")
            print(f"  Request Throughput: {m.request_throughput:.2f} req/sec | Avg embed time: {m.avg_embed_ms:.0f}ms (concurrent: {m.concurrent_users})")
    
    def _display_summary(self):
        """Display final summary"""
        if RICH_AVAILABLE:
            console.print("\n[bold cyan]📈 Test Summary - Phase Analysis[/bold cyan]\n")
            
            table = Table(title="Performance by Load Level")
            table.add_column("Users", justify="right")
            table.add_column("Embed", justify="right")
            table.add_column("Search", justify="right")
            table.add_column("MMR", justify="right")
            table.add_column("Cross", justify="right")
            table.add_column("Traverse", justify="right")
            table.add_column("Total", justify="right")
            table.add_column("Req/s", justify="right", style="cyan")
            table.add_column("Cause", style="magenta")
            table.add_column("Bottleneck", style="yellow")
            
            for m in self.step_metrics:
                table.add_row(
                    str(m.concurrent_users),
                    f"{m.avg_embed_ms:.0f}",
                    f"{m.avg_hybrid_ms:.0f}",
                    f"{m.avg_mmr_ms:.0f}",
                    f"{m.avg_cross_doc_ms:.0f}",
                    f"{m.avg_graph_traverse_ms:.0f}",
                    f"{m.avg_total_request_ms:.0f}",
                    f"{m.request_throughput:.1f}",
                    m.cause_label,
                    m.slowest_phase
                )
            
            console.print(table)
        else:
            print("\n📈 Test Summary - Phase Analysis\n")
            for m in self.step_metrics:
                print(f"Users: {m.concurrent_users} | Embed: {m.avg_embed_ms:.0f}ms | "
                      f"Search: {m.avg_hybrid_ms:.0f}ms | MMR: {m.avg_mmr_ms:.0f}ms | "
                      f"Cross: {m.avg_cross_doc_ms:.0f}ms | Traverse: {m.avg_graph_traverse_ms:.0f}ms | "
                      f"Total: {m.avg_total_request_ms:.0f}ms | Req/s: {m.request_throughput:.1f} | "
                      f"Bottleneck: {m.slowest_phase}")
    
    def _save_results(self):
        """Save results to Excel with rich formatting and charts"""
        if RICH_AVAILABLE:
            console.print("\n[cyan]💾 Saving results with rich formatting...[/cyan]")
        else:
            print("\n💾 Saving results with rich formatting...")

        self.config.OUTPUT_DIR.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.config.OUTPUT_DIR / f"neo4j_phase_profile_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                # Create workbook and get active sheet
                workbook = writer.book

                if EXCEL_STYLING_AVAILABLE:
                    # Define styles
                    header_style = NamedStyle(name="header_style")
                    header_style.font = Font(bold=True, color="FFFFFF")
                    header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_style.alignment = Alignment(horizontal="center", vertical="center")
                    header_style.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    data_style = NamedStyle(name="data_style")
                    data_style.alignment = Alignment(horizontal="right")
                    data_style.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    bottleneck_style = NamedStyle(name="bottleneck_style")
                    bottleneck_style.font = Font(color="FF0000", bold=True)
                    bottleneck_style.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    bottleneck_style.alignment = Alignment(horizontal="center")
                    bottleneck_style.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                # ===== DASHBOARD SHEET =====
                dashboard_df = self._create_dashboard_data()
                dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=1)

                worksheet = writer.sheets["Dashboard"]
                worksheet["A1"] = "Neo4j Phase Profiler - Results Dashboard"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=16, bold=True, color="366092")

                    # Apply styles to dashboard
                    for row in worksheet.iter_rows(min_row=2, max_row=len(dashboard_df)+1, min_col=1, max_col=len(dashboard_df.columns)):
                        for cell in row:
                            if cell.row == 2:  # Header row
                                cell.style = header_style
                            else:
                                cell.style = data_style

                # ===== PHASE BREAKDOWN SHEET =====
                breakdown_df = self._create_phase_breakdown_data()
                breakdown_df.to_excel(writer, sheet_name="Phase_Breakdown", index=False, startrow=1)

                worksheet = writer.sheets["Phase_Breakdown"]
                worksheet["A1"] = "Phase Timing Breakdown by Load Level"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=14, bold=True, color="366092")

                    # Apply styles and conditional formatting
                    for row in worksheet.iter_rows(min_row=2, max_row=len(breakdown_df)+1, min_col=1, max_col=len(breakdown_df.columns)):
                        for cell in row:
                            if cell.row == 2:  # Header row
                                cell.style = header_style
                            else:
                                cell.style = data_style

                    # Highlight bottleneck phases
                    bottleneck_col = None
                    for col_num, col_name in enumerate(breakdown_df.columns, 1):
                        if col_name == "Bottleneck":
                            bottleneck_col = col_num
                            break

                    if bottleneck_col:
                        for row in range(3, len(breakdown_df)+2):  # Data rows
                            cell = worksheet.cell(row=row, column=bottleneck_col)
                            cell.style = bottleneck_style

                # ===== PERFORMANCE TREND SHEET =====
                trend_df = self._create_performance_trend_data()
                trend_df.to_excel(writer, sheet_name="Performance_Trend", index=False, startrow=1)

                worksheet = writer.sheets["Performance_Trend"]
                worksheet["A1"] = "Performance Trends by Concurrent Users"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=14, bold=True, color="366092")

                    # Apply styles
                    for row in worksheet.iter_rows(min_row=2, max_row=len(trend_df)+1, min_col=1, max_col=len(trend_df.columns)):
                        for cell in row:
                            if cell.row == 2:  # Header row
                                cell.style = header_style
                            else:
                                cell.style = data_style

                # ===== DETAILED RESULTS SHEET =====
                detailed_df = pd.DataFrame([r.to_dict() for r in self.all_results])
                detailed_df.to_excel(writer, sheet_name="Detailed_Results", index=False, startrow=1)

                worksheet = writer.sheets["Detailed_Results"]
                worksheet["A1"] = "Individual Request Details"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=14, bold=True, color="366092")

                    # Apply styles to detailed results
                    for row in worksheet.iter_rows(min_row=2, max_row=len(detailed_df)+1, min_col=1, max_col=len(detailed_df.columns)):
                        for cell in row:
                            if cell.row == 2:  # Header row
                                cell.style = header_style
                            else:
                                cell.style = data_style

                # ===== RAW METRICS SHEET =====
                summary_df = pd.DataFrame([m.to_dict() for m in self.step_metrics])
                summary_df.to_excel(writer, sheet_name="Raw_Metrics", index=False, startrow=1)

                # ===== SYSTEM METRICS SHEET =====
                system_df = self._create_system_metrics_data()
                system_df.to_excel(writer, sheet_name="System_Metrics", index=False, startrow=1)
                # ===== EMBEDDING DETAIL SHEET =====
                embed_detail_df = self._create_embedding_detail_data()
                if not embed_detail_df.empty:
                    embed_detail_df.to_excel(writer, sheet_name="Embedding_Detail", index=False, startrow=1)
                    worksheet = writer.sheets["Embedding_Detail"]
                    worksheet["A1"] = "Embedding Sub-Phase Timings"
                    if EXCEL_STYLING_AVAILABLE:
                        worksheet["A1"].font = Font(size=14, bold=True, color="366092")
                        for row in worksheet.iter_rows(min_row=2, max_row=len(embed_detail_df)+1, min_col=1, max_col=len(embed_detail_df.columns)):
                            for cell in row:
                                if cell.row == 2:
                                    cell.style = header_style
                                else:
                                    cell.style = data_style

                worksheet = writer.sheets["System_Metrics"]
                worksheet["A1"] = "System Metrics by Load Level"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=14, bold=True, color="366092")

                    for row in worksheet.iter_rows(min_row=2, max_row=len(system_df)+1, min_col=1, max_col=len(system_df.columns)):
                        for cell in row:
                            if cell.row == 2:
                                cell.style = header_style
                            else:
                                cell.style = data_style

                worksheet = writer.sheets["Raw_Metrics"]
                worksheet["A1"] = "Raw Performance Metrics"
                if EXCEL_STYLING_AVAILABLE:
                    worksheet["A1"].font = Font(size=14, bold=True, color="366092")

                    # Apply styles
                    for row in worksheet.iter_rows(min_row=2, max_row=len(summary_df)+1, min_col=1, max_col=len(summary_df.columns)):
                        for cell in row:
                            if cell.row == 2:  # Header row
                                cell.style = header_style
                            else:
                                cell.style = data_style

                # Auto-adjust column widths for all sheets
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            if RICH_AVAILABLE:
                console.print(f"[green]✓ Rich Excel report saved to: {filename}[/green]")
                if EXCEL_STYLING_AVAILABLE:
                    console.print(f"[dim]📊 Includes: Dashboard, Phase Breakdown, Performance Trends, Detailed Results, Raw Metrics[/dim]")
                    console.print(f"[dim]🎨 Features: Color coding, borders, auto-sized columns, bottleneck highlighting[/dim]")
                else:
                    console.print(f"[dim]📊 Includes: Dashboard, Phase Breakdown, Performance Trends, Detailed Results, Raw Metrics[/dim]")
                    console.print(f"[yellow]⚠️  Install openpyxl for enhanced styling: pip install openpyxl[/yellow]")
                console.print(f"[dim]🧰 Added: System_Metrics sheet for CPU/Mem/Load/IO per step[/dim]")
            else:
                print(f"✓ Rich Excel report saved to: {filename}")
                print("📊 Includes: Dashboard, Phase Breakdown, Performance Trends, Detailed Results, Raw Metrics, System_Metrics")

        except Exception as e:
            if RICH_AVAILABLE:
                console.print(f"[red]✗ Error saving Excel: {e}[/red]")
            else:
                print(f"✗ Error saving Excel: {e}")

    def _create_dashboard_data(self) -> pd.DataFrame:
        """Create dashboard summary data"""
        if not self.step_metrics:
            return pd.DataFrame()

        latest = self.step_metrics[-1]  # Most recent results

        dashboard_data = {
            "Metric": [
                "Test Configuration",
                "Concurrent Users Tested",
                "Total Requests",
                "Success Rate",
                "Average Response Time",
                "P95 Response Time",
                "Bottleneck Phase",
                    "Bottleneck Share",
                    "Likely Cause",
                    "Top Contributors",
                "Embedding Time (Avg)",
                "Hybrid Search (Avg)",
                "MMR (Avg)",
                "Cross Doc (Avg)",
                "Graph Traverse (Avg)",
                "Format Context (Avg)",
                "Host CPU Avg",
                "CPU iowait Avg",
                "Memory Avg",
                "Load1/Core Avg",
                "Net RX/TX (MB/s)",
                "Disk R/W (MB/s)",
                    "Client JSON Parse (Avg)",
                    "Response Size (Avg KB)",
                "Test Duration",
                "Timestamp"
            ],
            "Value": [
                f"{self.config.MIN_USERS}-{self.config.MAX_USERS} users",
                latest.concurrent_users,
                sum(m.total_requests for m in self.step_metrics),
                f"{latest.success_rate:.1f}%",
                f"{latest.avg_total_request_ms:.1f}ms",
                f"{latest.p95_total_request_ms:.1f}ms",
                latest.slowest_phase.replace("_", " ").title(),
                    f"{latest.slowest_phase_pct:.1f}%",
                    latest.cause_label.replace("_", " "),
                latest.top_processing_contributors,
                f"{latest.avg_embed_ms:.1f}ms",
                f"{latest.avg_hybrid_ms:.1f}ms",
                f"{latest.avg_mmr_ms:.1f}ms",
                f"{latest.avg_cross_doc_ms:.1f}ms",
                f"{latest.avg_graph_traverse_ms:.1f}ms",
                f"{latest.avg_format_context_ms:.1f}ms",
                f"{latest.avg_cpu_percent:.1f}%",
                f"{latest.avg_cpu_iowait_percent:.1f}%",
                f"{latest.avg_mem_percent:.1f}%",
                f"{latest.load1_per_core_avg:.2f}",
                f"{latest.net_rx_mb_s:.2f}/{latest.net_tx_mb_s:.2f}",
                f"{latest.disk_read_mb_s:.2f}/{latest.disk_write_mb_s:.2f}",
                    f"{latest.avg_client_json_parse_ms:.1f}ms",
                    f"{(latest.avg_response_bytes/1024):.1f} KB",
                f"{len(self.step_metrics)} load steps",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }

        return pd.DataFrame(dashboard_data)

    def _create_phase_breakdown_data(self) -> pd.DataFrame:
        """Create phase breakdown table matching console output"""
        breakdown_data = []

        for m in self.step_metrics:
            total = m.avg_total_request_ms if m.avg_total_request_ms > 0 else 1
            
            # Calculate overhead (unmeasured time)
            total_tracked = (m.avg_embed_ms + m.avg_hybrid_ms + m.avg_mmr_ms + 
                           m.avg_cross_doc_ms + m.avg_graph_traverse_ms + m.avg_format_context_ms)
            overhead_ms = total - total_tracked
            
            # Calculate and round all percentages
            embed_pct = round(m.avg_embed_ms / total * 100, 1)
            hybrid_pct = round(m.avg_hybrid_ms / total * 100, 1)
            mmr_pct = round(m.avg_mmr_ms / total * 100, 1)
            cross_pct = round(m.avg_cross_doc_ms / total * 100, 1)
            traverse_pct = round(m.avg_graph_traverse_ms / total * 100, 1)
            format_pct = round(m.avg_format_context_ms / total * 100, 1)
            
            # Calculate overhead percentage to force sum to exactly 100%
            percentage_sum = embed_pct + hybrid_pct + mmr_pct + cross_pct + traverse_pct + format_pct
            overhead_pct = round(100.0 - percentage_sum, 1)

            row = {
                "Concurrent Users": m.concurrent_users,
                "Embedding (ms)": round(m.avg_embed_ms, 1),
                "Embedding (%)": embed_pct,
                "Hybrid Search (ms)": round(m.avg_hybrid_ms, 1),
                "Hybrid Search (%)": hybrid_pct,
                "MMR (ms)": round(m.avg_mmr_ms, 1),
                "MMR (%)": mmr_pct,
                "Cross Doc (ms)": round(m.avg_cross_doc_ms, 1),
                "Cross Doc (%)": cross_pct,
                "Graph Traverse (ms)": round(m.avg_graph_traverse_ms, 1),
                "Graph Traverse (%)": traverse_pct,
                "Format Context (ms)": round(m.avg_format_context_ms, 1),
                "Format Context (%)": format_pct,
                "Other/Overhead (ms)": round(overhead_ms, 1),
                "Other/Overhead (%)": overhead_pct,
                "Total Request (ms)": round(m.avg_total_request_ms, 1),
                "Bottleneck": m.slowest_phase.replace("_", " ").title()
            }
            breakdown_data.append(row)

        return pd.DataFrame(breakdown_data)

    def _create_embedding_detail_data(self) -> pd.DataFrame:
        """Create embedding sub-phase breakdown if available"""
        rows = []
        for r in self.all_results:
            if not r.success:
                continue
            p = r.phases
            if any(getattr(p, k, None) for k in [
                'embed_queue_wait_ms','embed_tokenize_ms','embed_forward_ms','embed_postprocess_ms'
            ]):
                rows.append({
                    "Step": r.step,
                    "User": r.user_index,
                    "Concurrent": r.concurrent_embeds,
                    "Queue Wait (ms)": round((p.embed_queue_wait_ms or 0), 1),
                    "Tokenize (ms)": round((p.embed_tokenize_ms or 0), 1),
                    "Forward (ms)": round((p.embed_forward_ms or 0), 1),
                    "Postprocess (ms)": round((p.embed_postprocess_ms or 0), 1),
                    "Embed Total (ms)": round((p.embed_time_ms or 0), 1)
                })
        return pd.DataFrame(rows)

    def _create_performance_trend_data(self) -> pd.DataFrame:
        """Create performance trend data for charting"""
        trend_data = []

        for m in self.step_metrics:
            row = {
                "Concurrent Users": m.concurrent_users,
                "Success Rate (%)": round(m.success_rate, 1),
                "Avg Response Time (ms)": round(m.avg_total_request_ms, 1),
                "P95 Response Time (ms)": round(m.p95_total_request_ms, 1),
                "Embedding Time (ms)": round(m.avg_embed_ms, 1),
                "Hybrid Search (ms)": round(m.avg_hybrid_ms, 1),
                "MMR (ms)": round(m.avg_mmr_ms, 1),
                "Cross Doc (ms)": round(m.avg_cross_doc_ms, 1),
                "Graph Traverse (ms)": round(m.avg_graph_traverse_ms, 1),
                "Format Context (ms)": round(m.avg_format_context_ms, 1),
                "Client JSON Parse (ms)": round(m.avg_client_json_parse_ms, 1),
                "Resp Size (KB)": round(m.avg_response_bytes / 1024, 1) if m.avg_response_bytes else 0,
                "Throughput (req/s)": round(m.throughput_rps, 2)
            }
            trend_data.append(row)

        return pd.DataFrame(trend_data)

    def _create_system_metrics_data(self) -> pd.DataFrame:
        """Create system metrics data per step for diagnosis"""
        rows = []
        for m in self.step_metrics:
            rows.append({
                "Concurrent Users": m.concurrent_users,
                "CPU Avg (%)": round(m.avg_cpu_percent, 1),
                "CPU iowait Avg (%)": round(m.avg_cpu_iowait_percent, 1),
                "Mem Avg (%)": round(m.avg_mem_percent, 1),
                "Swap Avg (%)": round(m.avg_swap_percent, 1),
                "Load1/Core Avg": round(m.load1_per_core_avg, 2),
                "Net RX (MB/s)": round(m.net_rx_mb_s, 2),
                "Net TX (MB/s)": round(m.net_tx_mb_s, 2),
                "Disk Read (MB/s)": round(m.disk_read_mb_s, 2),
                "Disk Write (MB/s)": round(m.disk_write_mb_s, 2),
                "Bottleneck": m.slowest_phase.replace("_", " ").title(),
                "Likely Cause": m.cause_label.replace("_", " ")
            })
        return pd.DataFrame(rows)


# ========================================
# MAIN
# ========================================

def main():
    """Entry point"""
    config = Config()
    config.OUTPUT_DIR.mkdir(exist_ok=True)
    
    profiler = PhaseProfiler(config)
    profiler.run_test()
    
    if RICH_AVAILABLE:
        console.print("\n[bold green]✓ Phase profiling completed![/bold green]\n")
    else:
        print("\n✓ Phase profiling completed!\n")


if __name__ == "__main__":
    main()